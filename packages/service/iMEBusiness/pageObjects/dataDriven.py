import time
import logging
import openpyxl
from selenium.webdriver.common.by import By
from iMEBusiness.pageObjects.BasePage import BasePage
from iMEBusiness.Utilities.excelReader import excel_Data
from iMEBusiness.Utilities.customLogger import Logger

log = Logger(__name__, logging.INFO)


class data_Driven(BasePage):

    def __init__(self, driver):
        super().__int__(driver)

    def data_driven_function(self):
        # path = openpyxl.load_workbook(".\\excel\\testData.xlsx")
        # sheet = path["data"]
        # row = sheet.max_row
        # col = sheet.max_column
        #
        # print(row)
        # print(col)
        path = ".\\excel\\testData.xlsx"
        sheetName = 'data'
        # row = excel_Data.getRowCount(".\\excel\\testData.xlsx", 'data')
        # log.logger.info("" + str(row))
        # col = excel_Data.getColCount(".\\excel\\testData.xlsx", 'data')
        # log.logger.info("" + str(col))
        row = excel_Data.getRowCount(path, sheetName)
        log.logger.info("" + str(row))
        col = excel_Data.getColCount(path, sheetName)
        log.logger.info("" + str(col))

        for i in range(2, row + 1):
            for j in range(1, col + 1):
                cellData = excel_Data.getCellData(path, sheetName, i, j)
                log.logger.info("" + str(cellData))
