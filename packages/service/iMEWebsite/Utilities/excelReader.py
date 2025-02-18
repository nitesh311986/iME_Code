import openpyxl


class excel_Data:
    @staticmethod
    def getRowCount(path, sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        return sheet.max_row

    @staticmethod
    def getColCount(path, sheetName):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        return sheet.max_column

    @staticmethod
    def getCellData(path, sheetName, rowNum, colNum):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        return sheet.cell(row=rowNum, column=colNum).value

    @staticmethod
    def setCellData(path, sheetName, rowNum, colNum, data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetName]
        sheet.cell(row=rowNum, column=colNum).value = data
        workbook.save(path)
